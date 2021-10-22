"""
 Openpyxl - Planilhas do Excel em Python
"""
import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
# planilha1 = pedidos['Página1']

# print(planilha1['b4'].value)
"""
for linha in planilha1['a1:c2']:
    for coluna in linha:
        print(coluna.value)
"""
"""
for linha in planilha1:
    for coluna in linha:
        print(coluna.value)
"""
"""for linha in planilha1:
    print(linha[0].value, linha[1].value, linha[2].value)
"""
"""for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)

    planilha1.cell(linha, 3).value = f'R$ {preco}'


pedidos.save('nova_planilha.xlsx')"""
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)

    planilha1.cell(linha, 3).value = f'R$ {preco}'

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha2.cell(linha, 1).value = numero_pedido
    planilha2.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)

    planilha2.cell(linha, 3).value = f'R$ {preco}'

planilha.save('new.xlsx')
